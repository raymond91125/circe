"""Ingest per-cell gene-expression datasets (a matrix of neuron class × gene marks).

The first such dataset is Cook 2020 Supplemental Data 6 (SI6): pharyngeal neuron classes ×
genes in four functional categories (metabotropic/ionotropic neurotransmitter receptors,
innexins, neuropeptides), marked "X"/"x" where a class expresses a gene. Genes are keyed to
persistent WormBase gene ids via a vendored, pre-resolved map (``si6_genes.csv``), so the build
is offline and deterministic. Class-level rows are expanded to individual cells by the build.

This module only parses; the build (assemble) maps class labels to cells and mints records.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path

import openpyxl

#: SI6 expression marks -> ExpressionConfidence enum values (x is undefined in the source).
_CONFIDENCE = {"X": "reported", "x": "putative"}


@dataclass(frozen=True)
class GeneRecord:
    wbgene: str  # e.g. "WB:WBGene00001517"
    symbol: str  # e.g. "gar-1"
    category: str  # GeneCategory enum value
    systematic_name: str


@dataclass(frozen=True)
class ExpressionRecord:
    class_label: str  # SI6 row label, e.g. "I1L/R"
    wbgene: str
    isoform: str  # "" when gene-level
    confidence: str  # "reported" / "putative"


@dataclass(frozen=True)
class GeneExpressionData:
    genes: list[GeneRecord]
    expressions: list[ExpressionRecord]


def load_gene_map(csv_path: Path) -> dict[str, dict[str, str]]:
    """si6 column label -> {symbol, isoform, wbgene, category, systematic_name}."""
    with open(csv_path, newline="") as f:
        return {row["si6_label"]: row for row in csv.DictReader(f)}


def read_gene_expression(xlsx_path: Path, gene_map_path: Path) -> GeneExpressionData:
    """Parse an SI6-shaped expression matrix + its resolved gene map."""
    gmap = load_gene_map(gene_map_path)

    # distinct Gene entities (keyed by WBGene; isoform columns collapse to one gene)
    genes: dict[str, GeneRecord] = {}
    for r in gmap.values():
        if r["wbgene"]:
            genes[r["wbgene"]] = GeneRecord(
                r["wbgene"], r["symbol"], r["category"], r["systematic_name"]
            )

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    name_row = rows[1]  # row 2: attribute headers + gene column names

    # columns whose header matches a gene in the map are gene columns
    gene_cols = {i: str(n).strip() for i, n in enumerate(name_row) if str(n).strip() in gmap}

    expressions: list[ExpressionRecord] = []
    for row in rows[2:]:
        label = row[0]
        if not label:
            continue
        label = str(label).strip()
        for col, si6_label in gene_cols.items():
            mark = row[col]
            if mark in _CONFIDENCE:
                r = gmap[si6_label]
                expressions.append(
                    ExpressionRecord(label, r["wbgene"], r["isoform"], _CONFIDENCE[mark])
                )
    return GeneExpressionData(genes=list(genes.values()), expressions=expressions)

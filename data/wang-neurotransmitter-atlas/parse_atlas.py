#!/usr/bin/env python
"""Parse the Wang et al. neurotransmitter atlas (Table S2) into per-neuron assignments.

Source: TableS2_expression.xlsx — supplement of Wang, Vidal, Sural et al., "A neurotransmitter
atlas of C. elegans males and hermaphrodites", eLife 2024 (doi:10.7554/eLife.95402). See
MANIFEST.md. The sheet has per-neuron reporter-expression columns (eat-4, unc-17, unc-25,
unc-47, cat-1, tph-1, cat-2, …) plus a consolidated "Neurotransmitter(s)" call; we take the
consolidated call (the reporter columns are color-coded and not machine-readable via values).

IMPORTANT: the "Neurotransmitter(s)" heading spans THREE merged sub-columns — U/V/W (primary /
secondary / tertiary transmitter). Reading only column U (as an earlier version did) drops the
co-transmitters that live in V/W (e.g. RIH's 5-HT uptake, RIM's tyramine + betaine, and the
cat-1/snf-3 betaine-uptake calls). We read all three and join the distinct parts with " | ".
Merged cells are resolved explicitly (read_only+values_only cannot see merge ranges).

Run:  uv run --with 'openpyxl' python parse_atlas.py
Output: assignments.csv (class, neuron, neurotransmitters) next to this script.
"""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
XLSX = HERE / "TableS2_expression.xlsx"
HEADER_ROW = 4  # 1-based; columns: Class | Neuron | Lineage | reporters… | Neurotransmitter(s)
COL_CLASS, COL_NEURON = 2, 3  # 1-based
COLS_NT = (21, 22, 23)  # 1-based U/V/W — the merged "Neurotransmitter(s)" block


def parse(path: Path = XLSX) -> list[dict]:
    ws = openpyxl.load_workbook(path, data_only=True).worksheets[0]

    def value(row: int, col: int) -> str:
        """Cell value, resolving merged ranges to their top-left anchor."""
        c = ws.cell(row=row, column=col)
        if c.value not in (None, ""):
            return str(c.value).strip()
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                anchor = ws.cell(row=rng.min_row, column=rng.min_col).value
                return str(anchor).strip() if anchor not in (None, "") else ""
        return ""

    out: list[dict] = []
    current_class = ""
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        neuron = value(row, COL_NEURON)
        if not neuron:
            continue
        current_class = value(row, COL_CLASS) or current_class  # class merged over members
        parts: list[str] = []
        for col in COLS_NT:
            part = value(row, col)
            if part and part not in parts:  # U:W is often merged to a single value
                parts.append(part)
        if parts:
            out.append(
                {"class": current_class, "neuron": neuron, "neurotransmitters": " | ".join(parts)}
            )
    return out


def main() -> None:
    records = parse()
    with (HERE / "assignments.csv").open("w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=["class", "neuron", "neurotransmitters"])
        w.writeheader()
        w.writerows(records)
    print(f"parsed {len(records)} neuron assignments -> assignments.csv")


if __name__ == "__main__":
    main()
